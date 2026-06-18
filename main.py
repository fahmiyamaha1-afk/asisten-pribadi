import disnake
from disnake.ext import commands, tasks
import json
import os
from datetime import datetime, timedelta

# Import pustaka pembuat Spreadsheet Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOKEN_BOT_KEUANGAN = os.getenv("TOKEN_KEUANGAN")

intents = disnake.Intents.default()
intents.message_content = True
intents.dm_messages = True

bot = commands.Bot(command_prefix="!", intents=intents)
FILE_DATA = "keuangan_pribadi.json"

def muat_data():
    if not os.path.exists(FILE_DATA):
        return {"masuk": 0, "keluar": 0, "riwayat_lengkap": [], "riwayat": []}
    try:
        with open(FILE_DATA, "r") as f:
            data = json.load(f)
            if "riwayat_lengkap" not in data:
                data["riwayat_lengkap"] = []
            return data
    except:
        return {"masuk": 0, "keluar": 0, "riwayat_lengkap": [], "riwayat": []}

def simpan_data(data):
    with open(FILE_DATA, "w") as f:
        json.dump(data, f, indent=4)

def format_rupiah(nominal):
    return f"Rp {nominal:,.0f}".replace(",", ".")

def buat_spreadsheet_laporan(data, nama_file="Laporan_Keuangan_30_Hari.xlsx"):
    """Fungsi khusus untuk merakit file Excel (.xlsx) yang rapi dan estetik"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Keuangan"
    
    # Aktifkan gridlines (garis kotak-kotak Excel)
    ws.views.sheetView[0].showGridLines = True
    
    # Pengaturan Gaya (Styles)
    font_judul = Font(name="Arial", size=16, bold=True, color="1A365D")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=11, bold=True)
    font_reguler = Font(name="Arial", size=11)
    
    fill_header = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    fill_sub_header = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_zebra = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    border_tipis = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0')
    )
    
    # 1. Judul Atas
    ws["A1"] = "LAPORAN BULANAN KEUANGAN PRIBADI"
    ws["A1"].font = font_judul
    ws["A2"] = f"Tanggal Cetak: {datetime.now().strftime('%d %B %Y')}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True)
    
    # 2. Kotak Ringkasan Saldo
    ws["A4"] = "📊 RINGKASAN SALDO"
    ws["A4"].font = font_bold
    
    ws["A5"] = "Total Pemasukan"
    ws["B5"] = data["masuk"]
    ws["A6"] = "Total Pengeluaran"
    ws["B6"] = data["keluar"]
    ws["A7"] = "Sisa Saldo Netto"
    ws["B7"] = "=B5-B6" # Menggunakan rumus formula Excel asli
    
    for r in range(5, 8):
        ws[f"A{r}"].font = font_bold if r == 7 else font_reguler
        ws[f"B{r}"].font = font_bold
        ws[f"B{r}"].number_format = '"Rp "#,##0'
        ws[f"A{r}"].fill = fill_sub_header if r == 7 else fill_zebra
        ws[f"B{r}"].fill = fill_sub_header if r == 7 else fill_zebra
        ws[f"A{r}"].border = border_tipis
        ws[f"B{r}"].border = border_tipis
        
    # 3. Tabel Detail Riwayat Transaksi Lengkap
    ws["A9"] = "📜 DAFTAR TRANSAKSI (30 HARI TERAKHIR)"
    ws["A9"].font = font_bold
    
    headers = ["Tanggal / Waktu", "Tipe Transaksi", "Nominal Transaksi"]
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_num)
        cell.value = header_title
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_tipis
        
    baris_mulai = 11
    if not data["riwayat_lengkap"]:
        ws.cell(row=baris_mulai, column=1, value="-").alignment = align_center
        ws.cell(row=baris_mulai, column=2, value="Belum ada data transaksi").font = font_reguler
        ws.cell(row=baris_mulai, column=3, value="-").alignment = align_center
        for c in range(1, 4):
            ws.cell(row=baris_mulai, column=c).border = border_tipis
    else:
        for idx, trx in enumerate(data["riwayat_lengkap"]):
            r_id = baris_mulai + idx
            
            c1 = ws.cell(row=r_id, column=1, value=trx['waktu'])
            c2 = ws.cell(row=r_id, column=2, value=trx['tipe'])
            c3 = ws.cell(row=r_id, column=3, value=trx['nominal'])
            
            c1.alignment = align_center
            c2.alignment = align_left
            c3.alignment = align_right
            
            c1.font = font_reguler
            c2.font = font_reguler
            c3.font = font_bold
            c3.number_format = '"Rp "#,##0' # Format Rupiah di Excel
            
            # Efek warna selang-seling (zebra) biar rapi dibaca
            if idx % 2 == 1:
                c1.fill = fill_zebra
                c2.fill = fill_zebra
                c3.fill = fill_zebra
                
            c1.border = border_tipis
            c2.border = border_tipis
            c3.border = border_tipis

    # Auto-fit lebar kolom agar teks tidak terpotong (###)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 4: # Lewati judul atas agar tidak kepanjangan
                continue
            if cell.value:
                # Jika valuenya angka rupiah, perkirakan panjang teksnya
                val_str = format_rupiah(cell.value) if isinstance(cell.value, (int, float)) else str(cell.value)
                if len(val_str) > max_len:
                    max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    wb.save(nama_file)
    return nama_file

@bot.event
async def on_ready():
    print("=============================================")
    print(f"✅ BOT KEUANGAN + MESIN SPREADSHEET EXCEL ONLINE!")
    print("=============================================")
    hitung_mundur_excel.start()

# ⏰ TIMER AUTOMATION: Ngecek berkala, kirim spreadsheet otomatis tiap 30 hari
@tasks.loop(hours=24)
async def hitung_mundur_excel():
    data = muat_data()
    
    if "terakhir_cetak" not in data:
        data["terakhir_cetak"] = datetime.now().strftime("%Y-%m-%d")
        simpan_data(data)
        return
        
    tgl_terakhir = datetime.strptime(data["terakhir_cetak"], "%Y-%m-%d")
    if datetime.now() >= tgl_terakhir + timedelta(days=30):
        nama_excel = buat_spreadsheet_laporan(data)
        
        if data.get("owner_id"):
            try:
                user = await bot.fetch_user(int(data["owner_id"]))
                await user.send(
                    f"📊 **LAPORAN REKAP SPREADSHEET BULANAN (30 HARI) PRIBADI AKANG**\n"
                    f"Berikut adalah file spreadsheet Excel (.xlsx) otomatis untuk memantau sirkulasi kas harian.",
                    file=disnake.File(nama_excel)
                )
                data["riwayat_lengkap"] = []
                data["terakhir_cetak"] = datetime.now().strftime("%Y-%m-%d")
                simpan_data(data)
            except Exception as e:
                print(f"Gagal mengirim Excel otomatis: {e}")
                
        if os.path.exists(nama_excel):
            os.remove(nama_excel)

@bot.event
async def on_message(message):
    if message.author.bot:
        return

    teks = message.content.lower().strip()
    data = muat_data()
    
    if "owner_id" not in data or data["owner_id"] != str(message.author.id):
        data["owner_id"] = str(message.author.id)
        simpan_data(data)
    
    if teks.startswith("uang masuk") or teks.startswith("uang keluar"):
        is_masuk = teks.startswith("uang masuk")
        bagian_teks = teks.replace("uang masuk", "").replace("uang keluar", "")
        angka_saja = "".join([char for char in bagian_teks if char.isdigit()])
        
        if not angka_saja:
            await message.reply("❌ **Format salah, Kang!**\nContoh: `uang masuk 500.000`")
            return
            
        nominal = int(angka_saja)
        waktu_sekarang = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        if is_masuk:
            data["masuk"] += nominal
            tipe = "🟢 UANG MASUK"
            keterangan = "Catatan pemasukan berhasil disimpan."
        else:
            data["keluar"] += nominal
            tipe = "🔴 UANG KELUAR"
            keterangan = "Catatan pengeluaran berhasil dipotong."
            
        saldo_total = data["masuk"] - data["keluar"]
        
        data["riwayat"].append(f"{tipe}: +{format_rupiah(nominal)}" if is_masuk else f"{tipe}: -{format_rupiah(nominal)}")
        if len(data["riwayat"]) > 5:
            data["riwayat"].pop(0)
            
        data["riwayat_lengkap"].append({
            "waktu": waktu_sekarang,
            "tipe": "Pemasukan" if is_masuk else "Pengeluaran",
            "nominal": nominal
        })
        
        simpan_data(data)
        
        embed = disnake.Embed(
            title="📊 LAPORAN KAS PRIBADI",
            description=f"**{keterangan}**",
            color=disnake.Color.green() if is_masuk else disnake.Color.red()
        )
        embed.add_field(name="💰 Nominal Transaksi", value=f"**{format_rupiah(nominal)}**", inline=False)
        embed.add_field(name="📈 Total Pemasukan", value=format_rupiah(data["masuk"]), inline=True)
        embed.add_field(name="📉 Total Pengeluaran", value=format_rupiah(data["keluar"]), inline=True)
        embed.add_field(name="💳 TOTAL SALDO AKHIR", value=f"**{format_rupiah(saldo_total)}**", inline=False)
        
        if data["riwayat"]:
            embed.add_field(name="📜 Riwayat Terakhir", value="\n".join(data["riwayat"]), inline=False)
            
        embed.set_footer(text=f"Dicatat privat untuk: {message.author.name}")
        await message.reply(embed=embed)
        
    elif teks == "!keuangan":
        saldo_total = data["masuk"] - data["keluar"]
        embed = disnake.Embed(title="💳 TOTAL ISI BRANKAS SAAT INI", color=disnake.Color.blue())
        embed.add_field(name="📈 Total Pemasukan", value=format_rupiah(data["masuk"]), inline=True)
        embed.add_field(name="📉 Total Pengeluaran", value=format_rupiah(data["keluar"]), inline=True)
        embed.add_field(name="💰 TOTAL SALDO NETTO", value=f"**{format_rupiah(saldo_total)}**", inline=False)
        await message.reply(embed=embed)

    # 📌 PERINTAH CETAK EXCEL INSTAN MANUAL: Ketik ini kapan saja di DM untuk minta file Excel langsung
    elif teks == "!cetak-excel":
        await message.reply("⏳ *Sedang merakit data pembukuan ke file Spreadsheet Excel, mohon tunggu...*")
        nama_excel = buat_spreadsheet_laporan(data)
        await message.reply(
            f"✅ **Laporan Spreadsheet Berhasil Dibuat!**",
            file=disnake.File(nama_excel)
        )
        if os.path.exists(nama_excel):
            os.remove(nama_excel)
        
    elif teks == "!reset-keuangan":
        data = {"masuk": 0, "keluar": 0, "riwayat_lengkap": [], "riwayat": [], "terakhir_cetak": datetime.now().strftime("%Y-%m-%d")}
        simpan_data(data)
        await message.reply("🔄 **Buku kas pribadi berhasil dikosongkan kembali dari Rp 0!**")

    await bot.process_commands(message)

if TOKEN_BOT_KEUANGAN:
    bot.run(TOKEN_BOT_KEUANGAN)
else:
    print("❌ ERROR: Token tidak ditemukan di Variables Railway!")
